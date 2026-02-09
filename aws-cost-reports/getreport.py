#!/usr/bin/env python3
import subprocess
import json
import csv
from datetime import datetime, timedelta
import calendar
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def get_previous_month_dates():
    """Get dates for the previous month for AWS Cost Explorer"""
    today = datetime.now()

    # First day of current month (e.g., 2026-02-01)
    first_day_current = today.replace(day=1)

    # Last day of previous month (e.g., 2026-01-31)
    last_day_previous = first_day_current - timedelta(days=1)

    # First day of previous month (e.g., 2026-01-01)
    first_day_previous = last_day_previous.replace(day=1)

    # AWS Cost Explorer requires NEXT day as End date to include full previous month
    # Start=2026-01-01, End=2026-02-01 includes all of January up to 23:59:59 on Jan 31
    start_date = first_day_previous.strftime('%Y-%m-%d')
    end_date = first_day_current.strftime('%Y-%m-%d')  # First day of current month

    return start_date, end_date

def get_cost_by_tag():
    """Get cost data by cost-usage tag"""
    start_date, end_date = get_previous_month_dates()

    print(f"Retrieving data for tag 'cost-usage' for period: {start_date} - {end_date}")

    # Query to get data by tag
    cmd = [
        'aws', 'ce', 'get-cost-and-usage',
        '--time-period', f'Start={start_date},End={end_date}',
        '--granularity', 'DAILY',
        '--metrics', 'BlendedCost',
        '--group-by', 'Type=TAG,Key=cost-usage',
        '--output', 'json'
    ]

    try:
        result = subprocess.run(cmd, check=True, capture_output=True, text=True)
        return json.loads(result.stdout), start_date, end_date
    except subprocess.CalledProcessError as e:
        print(f"AWS CLI error: {e}")
        print(f"Stderr: {e.stderr}")
        return None, start_date, end_date
    except json.JSONDecodeError as e:
        print(f"JSON decoding error: {e}")
        return None, start_date, end_date

def generate_reports(data, start_date, end_date):
    """Generate CSV and XLS reports"""
    if not data or 'ResultsByTime' not in data:
        print("No data available for report generation")
        return

    # Collect data
    tag_values = set()
    dates = []
    cost_data = {}

    # Process AWS data
    for result in data['ResultsByTime']:
        date = result['TimePeriod']['Start']
        dates.append(date)

        if 'Groups' in result:
            for group in result['Groups']:
                tag_value = group['Keys'][0] if group['Keys'] else 'without tag'
                tag_values.add(tag_value)

                if tag_value not in cost_data:
                    cost_data[tag_value] = {}

                cost = group['Metrics']['BlendedCost']['Amount']
                cost_data[tag_value][date] = cost

    # Sort data
    dates.sort()
    tag_values = sorted(tag_values)

    # Generate CSV
    generate_csv(tag_values, dates, cost_data)

    # Generate XLS
    generate_xls(tag_values, dates, cost_data, start_date, end_date)

def generate_csv(tag_values, dates, cost_data):
    """Generate CSV file"""
    filename = 'costs.csv'

    with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
        writer = csv.writer(csvfile)

        # Header: tags in first column, dates in subsequent columns
        header = ['cost-usage'] + dates + ['Total costs($)']
        writer.writerow(header)

        # Data for each tag
        tag_totals = {}
        for tag in tag_values:
            row = [tag]
            tag_total = 0.0

            for date in dates:
                cost = cost_data.get(tag, {}).get(date, "")
                if cost:
                    cost = f"{float(cost):.2f}"
                row.append(cost)
                try:
                    if cost:
                        tag_total += float(cost)
                except ValueError:
                    pass

            row.append(f"{tag_total:.2f}")
            writer.writerow(row)
            tag_totals[tag] = tag_total

        # Total row
        total_row = ['cost-usage total']
        grand_total = 0.0

        for date in dates:
            date_total = 0.0
            for tag in tag_values:
                cost = cost_data.get(tag, {}).get(date, "")
                try:
                    if cost:
                        date_total += float(cost)
                except ValueError:
                    pass
            total_row.append(f"{date_total:.2f}" if date_total > 0 else "")
            grand_total += date_total

        total_row.append(f"{grand_total:.2f}")
        writer.writerow(total_row)

    print(f"CSV report saved to: {filename}")

def generate_xls(tag_values, dates, cost_data, start_date, end_date):
    """Generate XLS file with formatting"""
    filename = 'costs.xlsx'

    wb = Workbook()
    ws = wb.active
    ws.title = "Cost Report"

    # Styles
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    tag_font = Font(bold=True, size=11)
    tag_fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
    total_font = Font(bold=True, color="FF0000", size=11)
    total_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    border = Border(left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'))

    # Header
    ws['A1'] = 'cost-usage'
    for col, date in enumerate(dates, 2):
        ws.cell(row=1, column=col, value=date)
    ws.cell(row=1, column=len(dates)+2, value='Total costs($)')

    # Format header
    for col in range(1, len(dates) + 3):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Tag data
    for row_idx, tag in enumerate(tag_values, 2):
        # Tag in first column
        ws.cell(row=row_idx, column=1, value=tag)
        tag_cell = ws.cell(row=row_idx, column=1)
        tag_cell.font = tag_font
        tag_cell.fill = tag_fill
        tag_cell.border = border

        tag_total = 0.0

        # Date data
        for col_idx, date in enumerate(dates, 2):
            cost = cost_data.get(tag, {}).get(date, "")
            if cost:
                cost_value = float(cost)
                cell = ws.cell(row=row_idx, column=col_idx, value=cost_value)
                cell.number_format = '0.00'
                tag_total += cost_value
            else:
                cell = ws.cell(row=row_idx, column=col_idx, value=0)
            cell.border = border
            cell.alignment = Alignment(horizontal='right', vertical='center')

        # Tag total
        total_cell = ws.cell(row=row_idx, column=len(dates)+2, value=tag_total)
        total_cell.number_format = '0.00'
        total_cell.border = border
        total_cell.alignment = Alignment(horizontal='right', vertical='center')

    # Total row
    total_row_idx = len(tag_values) + 2
    ws.cell(row=total_row_idx, column=1, value='cost-usage total')
    total_header_cell = ws.cell(row=total_row_idx, column=1)
    total_header_cell.font = total_font
    total_header_cell.fill = total_fill
    total_header_cell.border = border

    grand_total = 0.0

    for col_idx, date in enumerate(dates, 2):
        date_total = 0.0
        for tag in tag_values:
            cost = cost_data.get(tag, {}).get(date, "")
            try:
                if cost:
                    date_total += float(cost)
            except ValueError:
                pass

        cell = ws.cell(row=total_row_idx, column=col_idx, value=date_total)
        cell.number_format = '0.00'
        cell.font = total_font
        cell.fill = total_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='right', vertical='center')

        grand_total += date_total

    # Grand total
    grand_total_cell = ws.cell(row=total_row_idx, column=len(dates)+2, value=grand_total)
    grand_total_cell.number_format = '0.00'
    grand_total_cell.font = total_font
    grand_total_cell.fill = total_fill
    grand_total_cell.border = border
    grand_total_cell.alignment = Alignment(horizontal='right', vertical='center')

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min((max_length + 2) * 1.2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Freeze panes
    ws.freeze_panes = 'B2'

    # === ADD FLOATING SHAPE ANNOTATION ===
    try:
        # Import required modules for Shape
        from openpyxl.drawing.spreadsheet_drawing import (
            AnchorMarker, OneCellAnchor, SpreadsheetDrawing
        )
        from openpyxl.drawing.xdr import XDRPositiveSize2D
        from openpyxl.drawing.text import (
            RichText, Paragraph, ParagraphProperties, CharacterProperties
        )
        from openpyxl.drawing.fill import SolidFill
        from openpyxl.drawing.line import LineProperties

        # Create drawing object
        drawing = SpreadsheetDrawing()

        # Position the Shape (floating text box)
        # Position: Starting at column C (2), row 14 (13 in 0-based)
        marker = AnchorMarker(
            col=2,    # Column C (A=0, B=1, C=2)
            colOff=0, # Horizontal offset from column
            row=13,   # Row 14 (1=0, 2=1, ..., 14=13)
            rowOff=0  # Vertical offset from row
        )

        # Size of the Shape (width=~8cm, height=~3cm)
        size = XDRPositiveSize2D(cx=6000000, cy=1500000)

        # Create anchor
        anchor = OneCellAnchor(_from=marker, ext=size)

        # Create RichText content
        rich_text = RichText()

        # Paragraph 1: Header
        p1 = Paragraph()
        p1.rpPr = ParagraphProperties(
            defRPr=CharacterProperties(
                sz=1200,  # Font size 12pt
                b=True,   # Bold
                solidFill=SolidFill(srgbClr="000000")  # Black color
            )
        )
        p1.add_run("Annotation for \"cost-usage$\":")
        rich_text.add(p1)

        # Paragraph 2: Main text
        p2 = Paragraph()
        p2.rpPr = ParagraphProperties(
            defRPr=CharacterProperties(
                sz=1000,  # Font size 10pt
                solidFill=SolidFill(srgbClr="333333")  # Dark gray
            )
        )
        p2.add_run("This figure is the total cost for untagged AWS resources for the previous month, automatically generated by a scheduled workflow.")
        rich_text.add(p2)

        # Paragraph 3: Additional text
        p3 = Paragraph()
        p3.rpPr = ParagraphProperties(
            defRPr=CharacterProperties(
                sz=1000,
                solidFill=SolidFill(srgbClr="333333")
            )
        )
        p3.add_run("The amount includes a monthly TAX fee, which is charged against untagged resources on the 1st.")
        rich_text.add(p3)

        # Create Shape with properties
        from openpyxl.drawing.shape import Shape

        shape = Shape(
            anchor=anchor,
            rich_text=rich_text,
            shape_type='rect'  # Rectangle shape
        )

        # Set Shape properties (yellow background, blue border)
        shape.fill = SolidFill(srgbClr="FFFFCC")  # Light yellow background
        shape.line = LineProperties(w=9525, solidFill=SolidFill(srgbClr="0000FF"))  # Blue border

        # Add shape to drawing
        drawing.add(shape)

        # Add drawing to worksheet
        ws.add_drawing(drawing)

        print("✓ Added floating Shape annotation for 'cost-usage$'")

    except ImportError as e:
        print(f"Note: Could not create floating Shape: {e}")
        print("Adding annotation as formatted text box instead...")

        # Fallback: formatted text box
        annotation_row = total_row_idx + 2
        annotation_text = """Annotation for "cost-usage$":
        
This figure is the total cost for untagged AWS resources for the previous month, automatically generated by a scheduled workflow.
        
The amount includes a monthly TAX fee, which is charged against untagged resources on the 1st."""

        annotation_cell = ws.cell(row=annotation_row, column=1, value=annotation_text)
        annotation_cell.font = Font(name='Calibri', size=10, bold=True, color="000000")
        annotation_cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
        annotation_cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        annotation_cell.border = Border(
            left=Side(style='thin', color='3366FF'),
            right=Side(style='thin', color='3366FF'),
            top=Side(style='thin', color='3366FF'),
            bottom=Side(style='thin', color='3366FF')
        )

        ws.merge_cells(
            start_row=annotation_row,
            start_column=1,
            end_row=annotation_row + 3,
            end_column=4
        )

    # Add info sheet
    info_ws = wb.create_sheet("Info")
    info_ws['A1'] = 'Cost Report Information'
    info_ws['A1'].font = Font(bold=True, size=14)

    info_data = [
        ['Report Period:', f"{start_date} to {end_date}"],
        ['Generated:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
        ['Total Tags:', len(tag_values)],
        ['Total Days:', len(dates)],
        ['Data Source:', 'AWS Cost Explorer'],
        ['Group By:', 'Tag: cost-usage']
    ]

    for row, (key, value) in enumerate(info_data, 3):
        info_ws.cell(row=row, column=1, value=key).font = Font(bold=True)
        info_ws.cell(row=row, column=2, value=value)

    # Auto-adjust column widths for info sheet
    for column in info_ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        info_ws.column_dimensions[column_letter].width = max_length + 2

    wb.save(filename)
    print(f"XLS report saved to: {filename}")

def check_available_tags():
    """Check available tags"""
    start_date, end_date = get_previous_month_dates()

    cmd = [
        'aws', 'ce', 'get-tags',
        '--time-period', f'Start={start_date},End={end_date}',
        '--output', 'json'
    ]

    try:
        result = subprocess.run(cmd, check=True, capture_output=True, text=True)
        data = json.loads(result.stdout)

        print("Available tags:")
        for tag in sorted(data.get('Tags', [])):
            print(f"  - {tag}")

        return 'cost-usage' in data.get('Tags', [])
    except Exception as e:
        print(f"Error checking tags: {e}")
        return False

def main():
    """Main function"""
    print("Generating report for tag 'cost-usage'...")

    # Check available tags
    if not check_available_tags():
        print("Warning: tag 'cost-usage' might be missing")

    # Get data
    data, start_date, end_date = get_cost_by_tag()

    if data:
        generate_reports(data, start_date, end_date)
        print("Reports generated successfully!")
    else:
        print("Failed to retrieve data from AWS")

if __name__ == "__main__":
    main()